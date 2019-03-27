# -*- utf-8; -*-

import sys

from PyQt5 import QtWidgets, uic, QtSql, QtGui, QtCore
import openpyxl


class MainWindow(QtWidgets.QMainWindow):
    def __init__(self):
        super(MainWindow, self).__init__()
        uic.loadUi("mainform.ui", self)

        self.db = QtSql.QSqlDatabase.addDatabase('QMYSQL')
        self.db.setHostName("localhost")
        self.db.setUserName("root")
        self.db.setPassword("")

        self.pushButtonConnectDB.clicked.connect(self.clickChooseDB)
        self.pushButtonSaveAsSQL.clicked.connect(self.clickSaveSQL)
        self.pushButtonOpenSQL.clicked.connect(self.clickOpenSQL)
        self.pushButtonSaveAsExcel.clicked.connect(self.clickSaveExcel)
        self.pushButtonOpenExcelFile.clicked.connect(self.clickOpenExcel)
        self.comboBox.currentTextChanged.connect(self.changeTable)


    def __del__(self):
        self.db.close()

    def changeTable(self, item):
        self.curModel.setTable(item)
        self.curModel.select()

    def clickChooseDB(self):
        item = QtWidgets.QInputDialog.getText(self, 'Choose database',
                                              'Choose your database name:')
        if not item[1]:
            return

        self.db.close()
        self.db.setDatabaseName(item[0])
        self.db.open()
        if not self.db.isOpen():
            print("Error connecton")
            print(self.db.lastError().text())
            return
        print("Соединение установлено...")

        curTable = self.db.tables()[0]
        self.curModel = QtSql.QSqlTableModel(self, self.db)
        self.curModel.setTable(curTable)
        self.curModel.select()
        self.tableView.setModel(self.curModel)

        for item in self.db.tables():
            self.comboBox.addItem(item)

    def clickOpenSQL(self):
        item = QtWidgets.QFileDialog.getOpenFileName(self, 'Choose file', '', '*.sql')
        if not item[1]:
            return

        filePath = item[0]
        fSQL = open(filePath, 'r')
        query = QtSql.QSqlQuery(self.db)
        if not query.exec(fSQL.read()):
            print(query.lastError().text())
        else:
            print("Файл прочитан, все запросы выполнены")
        self.comboBox.clear()
        self.comboBox.addItems(self.db.tables())
        for item in self.db.tables():
            self.comboBox.addItem(item)
        fSQL.close()

    def clickSaveSQL(self):
        item = QtWidgets.QFileDialog.getSaveFileName(self, 'Choose file', '', '*.sql')
        if not item[1]:
            return

        filePath = item[0]
        commandCreateTable = "CREATE TABLE `" + self.comboBox.currentText() + "` ("
        rowCount = self.curModel.rowCount()
        colCount = self.curModel.columnCount()
        for col in range(colCount - 1):
            commandCreateTable += "`" + str(self.curModel.headerData(col, QtCore.Qt.Horizontal, QtCore.Qt.DisplayRole)) + "` VARCHAR(255),\n"
        commandCreateTable += "`" + str(self.curModel.headerData(colCount - 1, QtCore.Qt.Horizontal, QtCore.Qt.DisplayRole)) + "`VARCHAR(255));\n"

        commandCreateTable += "INSERT INTO `" + self.comboBox.currentText() + "` VALUES\n"
        for i in range(rowCount - 1):
            commandCreateTable += "("
            for j in range(colCount - 1):
                commandCreateTable += str(self.curModel.data(self.curModel.index(i, j))) + ", "
            commandCreateTable += str(self.curModel.data(self.curModel.index(i, colCount - 1))) + "),\n"
            print(commandCreateTable)
        commandCreateTable += "("
        for j in range(colCount - 1):
            commandCreateTable += str(self.curModel.data(self.curModel.index(rowCount - 1, j))) + ", "
        commandCreateTable += str(self.curModel.data(self.curModel.index(rowCount - 1, colCount - 1))) + ");"

        fSQL = open(filePath, 'w')
        fSQL.write(commandCreateTable)
        fSQL.close()

    def clickSaveExcel(self):
        item = QtWidgets.QFileDialog.getSaveFileName(self, 'Choose file', '', '*.xlsx')
        if not item[1]:
            return

        filePath = item[0]
        fExcel = openpyxl.Workbook()
        curRow = []
        sheet = fExcel.active
        for i in range(self.curModel.rowCount()):
            for j in range(self.curModel.columnCount()):
                curRow.append(self.curModel.data(self.curModel.index(i, j)))
            sheet.append(curRow)
            curRow = []
        fExcel.save(filePath)

    def clickOpenExcel(self):
        item = QtWidgets.QFileDialog.getOpenFileName(self, 'Choose file', '', '*.xlsx')
        if not item[1]:
            return

        filePath = item[0]
        fExcel = openpyxl.load_workbook(filePath)
        curSheet = fExcel[fExcel.sheetnames[0]]
        myModel = QtGui.QStandardItemModel(self)
        colCount = 0
        rowCount = 0
        for row in curSheet.rows:
            colCount = 0
            for col in row:
                colCount += 1
            rowCount += 1
        myModel.setRowCount(rowCount)
        myModel.setColumnCount(colCount)
        curRow = 0
        curCol = 0
        for row in curSheet.rows:
            for col in row:
                myModel.setData(myModel.index(curRow, curCol), col.value)
                curCol += 1
            curRow += 1
            curCol = 0
        self.curModel = myModel
        self.tableView.setModel(self.curModel)
        self.comboBox.clear()

        # self.comboBox.addItem(fExcel.sheetnames[0])

app = QtWidgets.QApplication(sys.argv)
widget = MainWindow()
widget.show()
sys.exit(app.exec_())